from openpyxl import load_workbook

class DoEcel :
    def get_data(self , file_name , sheet_name) :
        file = load_workbook(file_name)
        sheet = file(sheet_name)

        test_data = []
        new_tel = file['tel'].cell(1,1).value

        for i in range(2, sheet.max_row+1) :
            sub_data = {}
            sub_data["id"] = sheet.cell(i,1).value
            sub_data['HttpMethod'] = sheet.cell(i,2).value
            sub_data['module'] = sheet.cell(i,3).value
            sub_data['description'] = sheet.cell(i,4).value
            sub_data['url'] = sheet.cell(i,5).value
            
            sub_data['param'] = sheet.cell(i,6).value

            sub_data['ExpectedResult'] = sheet.cell(i,7).value
            test_data.append(sub_data)
            file['tel'].cell(1,1).value = new_tel + 1
            file.save(file_name)
        return test_data

        def write_back(self,file_name , sheet_name , row , ActualResult , TestResult) :
            wb = load_workbook(file_name) 
            sheet = wb[sheet_name]
            sheet.cell(row,8).value = ActualResult
            sheet.cell(row,9).value = TestResultwb.save(file_name)

if __name__ == '__main__' :
    test_data = DoExcel().get_data('api.xlsx','testcase')
    print('测试数据是：{}'.formate(test_data))

               
